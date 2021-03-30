#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from win32api import GetSystemMetrics
from tkinter import *
from tkinter import font
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

file = pathlib.Path('Admission Form Record.xlsx')
adminfile = pathlib.Path('Admin Record.xlsx')

if file.exists():
    pass

else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'NAME'
    sheet['B1'] = 'PHONE NO.'
    sheet['C1'] = 'EMAIL ID'
    sheet['D1'] = 'GENDER'
    sheet['E1'] = 'PRESENT ADD.'
    sheet['F1'] = 'PERMANENT ADD.'
    sheet['G1'] = 'FATHER\'S NAME'
    sheet['H1'] = 'FATHER\'S PHN. NO.'
    sheet['I1'] = 'MOTHER\'S NAME'
    sheet['J1'] = 'MOTHER\'S PHN NO.'
    sheet['K1'] = 'CLASS X SCORE'
    sheet['L1'] = 'PERCENTAGE'
    sheet['M1'] = 'BOARD'
    sheet['N1'] = 'CLASS XII SCORE'
    sheet['O1'] = 'PERCENTAGE'
    sheet['P1'] = 'BOARD'
    
    file.save('Admission Form Record.xlsx')

if adminfile.exists():
    pass

else:
    adminfile = Workbook()
    adminsheet = adminfile.active
    adminsheet['A1'] = 'ID'
    adminsheet['B1'] = 'PASSWORD'
    
    adminfile.save('Admin Record.xlsx')

def destroy():
    center_frame6.destroy()
    center_frame8.destroy()
    center_frame10.destroy()
    
def quit():
    canvas.delete('all')
    center_frame7.destroy()
    
    
def read():
    global i
    
    record = openpyxl.load_workbook('Admission Form Record.xlsx')
    mini_database = record.active

    names = mini_database.cell(row = i, column = 1)
    
    phones = mini_database.cell(row = i, column = 2)
    
    emails = mini_database.cell(row = i, column = 3)

    genders = mini_database.cell(row = i, column = 4)

    present_add = mini_database.cell(row = i, column = 5)

    permanent_add = mini_database.cell(row = i, column = 6)

    father = mini_database.cell(row = i, column = 7)

    father_phones = mini_database.cell(row = i, column = 8)

    mother = mini_database.cell(row = i, column = 9)

    mother_phones = mini_database.cell(row = i, column = 10)

    X_score = mini_database.cell(row = i, column = 11)

    X_percentage = mini_database.cell(row = i, column = 12)

    X_board = mini_database.cell(row = i, column = 13)

    XII_score = mini_database.cell(row = i, column = 14)
   
    XII_percentage = mini_database.cell(row = i, column = 15)
   
    XII_board = mini_database.cell(row = i, column = 16)
    
    global center_frame6
    center_frame6 = Frame(center_frame10)
    center_frame6.place(relx = 0.0, rely = 0.0, height = 470, width = 710, anchor = NW)
    
    upperframe1 = Frame(center_frame6)
    upperframe1.place(x = 0, y = 10, height = 100, width = 110)
    
    upperframe2 = Frame(center_frame6)
    upperframe2.place(x = 111, y = 0, height = 140, width = 590)
    
    global logo_s
    logo_s = ImageTk.PhotoImage(Image.open("logo_stcet.png"))
    
    label = Label(upperframe1,image = logo_s).pack()
    
    desired_font = font.Font(font = 'Times', size = 10, weight = 'NORMAL')
    
    e1 = Label(upperframe2, text = 'ST. THOMAS\' COLLEGE OF ENGINEERING AND TECHNOLOGY',font = desired_font)
    e1.pack(pady = 10)
    e2 = Label(upperframe2, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack()
    e3 = Label(upperframe2, text = 'Contact No.: (+033)2448-1081')
    e3.pack()
    e4 = Label(upperframe2, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack()
    e5 = Label(upperframe2, text = 'DETAILED INFORMATION', fg='blue',font=desired_font)
    e5.pack()
   
    label_std_info = Label(center_frame6, text='Student\'s Information',font = ("Verdana",10,"bold"))
    label_std_info.place(x=20,y=145)
    
    label_name = Label(center_frame6, text = "Name: ")
    label_name.place(x = 20, y = 175)
    Label_Name1 = Label(center_frame6, text = names.value)
    Label_Name1.place(x = 130, y = 175)
    
    
    
    label_phn = Label(center_frame6,text="Phone No.: ")
    label_phn.place(x=375,y=175)
    Label_phn1=Label(center_frame6, text=phones.value)
    Label_phn1.place(x=510,y=175)
    
    
    label_email = Label(center_frame6, text = "E-mail: ")
    label_email.place(x = 20, y = 205)
    Label_email1 = Label(center_frame6, bd = 2, text = emails.value)
    Label_email1.place(x = 130, y = 205)
    
    
    label_gender = Label(center_frame6, text = "Gender: ")
    label_gender.place(x = 375, y = 205)
    
    Label_gender1 = Label(center_frame6, text = genders.value)
    Label_gender1.place(x = 510, y = 205)
  
    label_present_add = Label(center_frame6, text = "Present Address: ")
    label_present_add.place(x = 20, y = 235)

    Label_present_add1 = Label(center_frame6, text = present_add.value)
    Label_present_add1.place(x = 130, y = 235)
    
    
    label_per_add = Label(center_frame6, text = "Permanent Address: ")
    label_per_add.place(x = 375, y = 235)
    
    Label_per_add1 = Label(center_frame6, text = permanent_add.value)
    Label_per_add1.place(x = 510, y = 235) 

    
    label_par_info = Label(center_frame6, text = 'Parent\'s Information', font = ("Verdana",10,"bold"))
    label_par_info.place(x = 20, y = 265)
     
    label_fat_name = Label(center_frame6, text = "Father's Name: ")
    label_fat_name.place(x = 20, y = 295)
    
    Label_fat_name1 = Label(center_frame6, text = father.value)
    Label_fat_name1.place(x = 130, y = 295)
    
    
    label_mot_name = Label(center_frame6, text = "Mother's Name: ")
    label_mot_name.place(x = 20, y = 325)
    
    Label_mot_name1 = Label(center_frame6, text = mother.value)
    Label_mot_name1.place(x = 130, y = 325)
    
    
    label_fat_phn = Label(center_frame6, text = "Father's Phone No.: ")
    label_fat_phn.place(x = 375, y = 295)
    
    Label_fat_phn1 = Label(center_frame6, text = father_phones.value)
    Label_fat_phn1.place(x = 510, y = 295)
    
    
    label_mot_phn = Label(center_frame6, text = "Mother's Phone No.: ")
    label_mot_phn.place(x = 375, y = 325)
      
    Label_mot_phn1 = Label(center_frame6, text = mother_phones.value)
    Label_mot_phn1.place(x = 510, y = 325)
    
    
    label_class_10 = Label(center_frame6, text = "Class X", font = ("Verdana",10,"bold"))
    label_class_10.place(x = 20, y = 355)
    
    label_total_10 = Label(center_frame6, text = "Total Score: ")
    label_total_10.place(x = 20, y = 385)
    
    Label_total_10_1 = Label(center_frame6, bd = 2, text = X_score.value)
    Label_total_10_1.place(x = 107, y = 385)
    
    
    label_percent_10 = Label(center_frame6, text = "Percentage: ")
    label_percent_10.place(x = 270, y = 385)
    
    Label_percent_10_1 = Label(center_frame6, text = X_percentage.value)
    Label_percent_10_1.place(x = 360, y = 385)
    
    
    label_board_10 = Label(center_frame6, text = "Board: ")
    label_board_10.place(x = 522, y = 385)
    
    label_board_10_1 = Label(center_frame6, text = X_board.value)
    label_board_10_1.place(x = 580, y = 385)
    
    label_class_12 = Label(center_frame6, text = "Class XII", font = ("Verdana",10,"bold"))
    label_class_12.place(x = 20, y = 415)
    
    label_total_12 = Label(center_frame6, text = "Total Score: ")
    label_total_12.place(x = 20, y = 445)
    
    label_total_12_1 = Label(center_frame6, text = XII_score.value)
    label_total_12_1.place(x = 107, y = 445)
    
    
    label_percent_12 = Label(center_frame6, text = "Percentage: ")
    label_percent_12.place(x = 270, y = 445)
    
    Label_percent_12_1 = Label(center_frame6, text = XII_percentage.value)
    Label_percent_12_1.place(x = 360, y = 445)
    
    
    label_board_12 = Label(center_frame6, text = "Board: ")
    label_board_12.place(x = 522, y = 445)
    
    
    Label_board_12_1 = Label(center_frame6, text = XII_board.value)
    Label_board_12_1.place(x = 580, y = 445)
    

def check(num):
    global i
    global center_frame6
    
    if num == 0:
        read()
        
    elif num == 1:
        if i >= 2 and i < max_row:
            if(center_frame6): center_frame6.destroy()
            i = i+1
            read()
        elif i >= max_row:
            messagebox.showwarning('Status','End of Records!')
            
    elif num == 2:
        if i > 2 and i <= max_row:
            if(center_frame6): center_frame6.destroy()
            i = i-1
            read()
        elif i <= 2:
            messagebox.showwarning('Status','End of Records!')
       
    
def detailed_statement():
    global center_frame10
    center_frame10 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame10.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    
    record = openpyxl.load_workbook('Admission Form Record.xlsx')
    mini_database = record.active
    
    global max_row
    global i
    max_row = mini_database.max_row
    max_column = mini_database.max_column
    
    if max_row == 1:
        messagebox.showwarning('Status','No records found!')
    
    else:
        i = 2
        
        global center_frame8
        center_frame8 = Frame(center_frame10, height = 86, width = 710)
        center_frame8.place(relx = 0.0, rely = 1.0, anchor = SW)
        
        button1 = Button(center_frame8, text = 'Previous', command = lambda: check(int(2)), font = ('Arial', 14),bg = 'SteelBlue', fg = 'white', activebackground = 'silver', activeforeground = 'black')
        button1.place(x = 160, y = 30)
        
        button2 = Button(center_frame8, text = 'Quit', command = destroy, font = ('Arial', 14), bg = 'black', fg = 'white', activebackground = 'silver', activeforeground = 'black')
        button2.place(x = 330, y = 30)
        
        button3 = Button(center_frame8, text = 'Next', command = lambda: check(int(1)), font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
        button3.place(x = 460, y = 30)
        
        check(int(0))

    
def brief_statement():
    global center_frame7
    center_frame7 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame7.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    scroll = ttk.Scrollbar(center_frame7,orient = 'vertical')
    scroll.pack(side = RIGHT, fill = 'y')  
    global canvas
    canvas = Canvas(center_frame7, yscrollcommand = scroll.set)
    canvas.pack(side = LEFT, fill = 'both', expand = 'yes')
    scroll.config(command = canvas.yview)
    canvas.bind_all('<Configure>', lambda e : canvas.configure(scrollregion = canvas.bbox('all')))
    global center_frame9
    center_frame9 = Frame(canvas)
    canvas.create_window((0,0), window = center_frame9, anchor = NW)
    
    studentname = []     
    phonenumber = []
    emailID = [] 
    classXII = []
    global mini_record
    mini_record = openpyxl.load_workbook('Admission Form Record.xlsx')
    global mini_database
    mini_database = mini_record.active
    
    row = mini_database.max_row
    column = mini_database.max_column
    
    for i in range(2, row + 1):
        names = mini_database.cell(row = i, column = 1)
        studentname.append(str(names.value))
        
        phones = mini_database.cell(row = i, column = 2)
        phonenumber.append(str(phones.value))
        
        emails = mini_database.cell(row = i, column = 3)
        emailID.append(str(emails.value))
        
        XIIscores = mini_database.cell(row = i, column = 15)
        classXII.append(str(XIIscores.value))
    
    upper_frame = Frame(center_frame9)
    upper_frame.pack()
    global logo_s
    logo_s = ImageTk.PhotoImage(Image.open("logo_stcet.png"))
    label = Label(upper_frame, image = logo_s).grid(row = 0,column = 0, padx = 10, pady = 10)
    
    desired_font = font.Font(font = 'Times', size = 10, weight = 'NORMAL')
    
    e1 = Label(upper_frame, text = 'ST. THOMAS\' COLLEGE OF ENGINEERING AND TECHNOLOGY',font = desired_font)
    e1.grid(row = 0, column = 1, padx = 10, pady = 10)
    e2 = Label(upper_frame, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.grid(row = 1, column = 1)
    e3 = Label(upper_frame, text = 'Contact No.: (+033)2448-1081')
    e3.grid(row = 2, column = 1)
    e4 = Label(upper_frame, text = 'Mail ID: stcet2000@gmail.com')
    e4.grid(row = 3, column = 1)
    e5 = Label(upper_frame, text = 'BRIEF STATEMENT', fg = 'blue',font = desired_font)
    e5.grid(row = 4, column = 1, pady = 5)
    lower_frame = Frame(center_frame9)
    lower_frame.pack(pady = 10)
    
    name_heading = Label(lower_frame, text = 'NAME', font = ('Times', 12, 'bold'))
    name_heading.grid(row = 0, column = 0, padx = 20)
    ph_heading = Label(lower_frame, text ='PHONE NO.', font = ('Times', 12, 'bold'))
    ph_heading.grid(row = 0, column = 1, padx = 20)
    mail_heading = Label(lower_frame, text ='EMAIL ID', font = ('Times', 12, 'bold'))
    mail_heading.grid(row = 0, column = 2, padx = 20)
    marks_heading = Label(lower_frame, text = 'XII PERCENTAGE', font = ('Times', 12, 'bold'))
    marks_heading.grid(row = 0, column = 3, padx = 20)   
    for i in range(0,len(studentname)):
        stuname=Label(lower_frame, text = studentname[i], font = ('Arial', 9))
        stuname.grid(row = i+1, column = 0, padx = 20, pady = 10)
        phn=Label(lower_frame, text = phonenumber[i], font = ('Arial', 9))
        phn.grid(row = i+1, column = 1, padx = 20, pady = 10)
        stumail=Label(lower_frame, text = emailID[i], font = ('Arial', 9))
        stumail.grid(row = i+1,column = 2, padx = 20, pady = 10)
        stumarks=Label(lower_frame, text = classXII[i], font = ('Arial', 9))
        stumarks.grid(row = i+1, column = 3, padx = 20, pady = 10)
        
    button1 = Button(center_frame9, text = 'Quit', command = quit, bg = 'Red', fg = 'white', font = ('Arial', 13), activebackground = 'salmon', activeforeground = 'black')
    button1.pack(pady = 20)
    

def frame5():
    global center_frame5
    global i
    center_frame5 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame5.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    
    label = Label(center_frame5, image = logo).pack(ipadx = 30, ipady = 30)

    e1 = Label(center_frame5, text = "ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font = 10)
    e1.pack(ipadx = 30)
    e2 = Label(center_frame5, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack(ipadx = 30)
    e3 = Label(center_frame5, text = 'Contact No.: (+033)2448-1081')
    e3.pack(ipadx = 30)
    e4 = Label(center_frame5, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack(ipadx = 30)
    e5 = Label(center_frame5, text = 'STUDENT REGISTRATION FORM ', fg = 'blue', font = 13)
    e5.pack(ipadx = 30, ipady = 10)
    
    button1 = Button(center_frame5, text = 'Brief Statement', command = brief_statement, font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button1.place(x = 170, y = 420)
    
    button2 = Button(center_frame5, text = 'Detailed Statement', command = detailed_statement, font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button2.place(x = 380, y = 420)
    
    button3 = Button(center_frame5, text = 'Back', command = center_frame5.destroy, font = ('Arial', 14), bg = 'black', fg = 'white', activebackground = 'silver',activeforeground = 'black')
    button3.place(x = 320, y = 480)
    
    
def submit():
    a = name.get()
    b = phn_no.get()
    c = email_id.get()
    e = present_add.get()
    f = per_add.get()
    g = fat_name.get()
    h = fat_phn.get()
    i = mot_name.get()
    j = mot_phn.get()
    k = class_10_scr.get()
    l = class_10_perc.get()
    m = c1.get()
    n = class_12_scr.get()
    o = class_12_perc.get()
    p = c2.get()
    
    file = openpyxl.load_workbook('Admission Form Record.xlsx')
    sheet = file.active
    sheet.cell(column = 1, row = sheet.max_row + 1, value = a)
    sheet.cell(column = 2, row = sheet.max_row, value = b)
    sheet.cell(column = 3, row = sheet.max_row, value = c)
    sheet.cell(column = 5, row = sheet.max_row, value = e)
    sheet.cell(column = 6, row = sheet.max_row, value = f)
    sheet.cell(column = 7, row = sheet.max_row, value = g)
    sheet.cell(column = 8, row = sheet.max_row, value = h)
    sheet.cell(column = 9, row = sheet.max_row, value = i)
    sheet.cell(column = 10, row = sheet.max_row, value = j)
    sheet.cell(column = 11, row = sheet.max_row, value = k)
    sheet.cell(column = 12, row = sheet.max_row, value = l)
    sheet.cell(column = 13, row = sheet.max_row, value = m)
    sheet.cell(column = 14, row = sheet.max_row, value = n)
    sheet.cell(column = 15, row = sheet.max_row, value = o)
    sheet.cell(column = 16, row = sheet.max_row, value = p)
    
    if var.get() == 1:
        sheet.cell(column = 4, row = sheet.max_row, value = 'Male')
    elif var.get() == 2:
        sheet.cell(column = 4, row = sheet.max_row, value = 'Female')
    else:
        sheet.cell(column = 4, row = sheet.max_row, value = 'Others')
    
    file.save('Admission Form Record.xlsx')
    name.set('')
    phn_no.set('')
    email_id.set('')
    var.set(0)
    present_add.set('')
    per_add.set('')
    fat_name.set('')
    fat_phn.set('')
    mot_name.set('')
    mot_phn.set('')
    class_10_scr.set('')
    class_10_perc.set('')
    c1.set('Select')
    class_12_scr.set('')
    class_12_perc.set('')
    c2.set('Select')

def popup():
    messagebox.showinfo('Status','Registration Successful !') 
    
    
def fieldcheck():
    p_10 = entry_percent_10.get()
    p_12 = entry_percent_12.get()
    email_list = ['@gmail.com','@rediffmail.com','@yahoo.com','@hotmail.com']
    n1 = entry_email.get()   
    flag = 0
    for i in email_list:
        if i in n1:
            flag = 1
            break
    n2 = entry_phn.get()
    n_fat = entry_fat_phn.get()
    n_mot = entry_mot_phn.get()
    if(entry_name.get() == ''):
        messagebox.showwarning('Error','Name field can not be empty!')
    elif(entry_present_add.get() == ''):
        messagebox.showwarning('Error','Present address field can not be empty!')
    elif(entry_per_add.get() == ''):
        messagebox.showwarning('Error','Permanent address field can not be empty!')
    elif(entry_total_10.get() == '' or entry_total_12.get() == ''):
        messagebox.showwarning('Error','Total score field can not be empty!')
    elif(entry_fat_name.get() == ''):
        messagebox.showwarning('Error','Father\'s Name field can not be empty!')
    elif(entry_mot_name.get() == ''):
        messagebox.showwarning('Error','Mother\'s Name field can not be empty!')
    elif(len(n2)!=10 or len(n_fat)!=10 or len(n_mot)!=10):
        messagebox.showwarning('Error','Please enter a valid phone number!')   
    elif(flag == 0):
        messagebox.showwarning('Error','Please enter a valid E-mail ID!')
    elif(var.get() == 0):
        messagebox.showwarning('Error','Please select a gender!')
    elif(c1.get() == 'Select' or c2.get() == 'Select'):
        messagebox.showwarning('Error','Please select a board!')
    elif((int(p_10) < 0 or int(p_10) > 100) or (int(p_12) < 0 or int(p_12) > 100)):
        messagebox.showwarning('Error','Please enter a valid percentage!')
    else:
        submit()
        popup()

def frame4():
    
    center_frame4 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame4.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    upperframe1 = Frame(center_frame4)
    upperframe1.place(x=0,y=10, height = 100, width=110)
    upperframe2 = Frame(center_frame4)
    upperframe2.place(x=111,y=0, height=140,width=590)
    global logo_s
    logo_s = ImageTk.PhotoImage(Image.open("logo_stcet.png"))
    label = Label(upperframe1,image = logo_s).pack()
    desired_font = font.Font(font='Times', size=10, weight='NORMAL')
    e1 = Label(upperframe2, text = 'ST. THOMAS\' COLLEGE OF ENGINEERING AND TECHNOLOGY',font = desired_font)
    e1.pack(pady = 10)
    e2 = Label(upperframe2, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack()
    e3 = Label(upperframe2, text = 'Contact No.: (+033)2448-1081')
    e3.pack()
    e4 = Label(upperframe2, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack()
    e5 = Label(upperframe2, text = 'STUDENT REGISTRATION FORM', fg = 'blue', font = desired_font)
    e5.pack()
    
    label_std_info = Label(center_frame4, text= 'Student\'s Information',font = ("Verdana", 10, "bold"))
    label_std_info.place(x = 20, y = 145)
    
    label_name = Label(center_frame4, text = "Name: ")
    label_name.place(x = 20, y = 175)
    global entry_name
    entry_name = Entry(center_frame4, width = 28, bd = 2, textvariable = name)
    entry_name.place(x = 130, y = 175)
    
    
    
    label_phn = Label(center_frame4, text = "Phone No.: ")
    label_phn.place(x = 375,y = 175)
    global entry_phn
    entry_phn = Entry(center_frame4, width = 28, bd = 2, textvariable = phn_no)
    entry_phn.place(x = 510, y = 175)
    
    
    label_email = Label(center_frame4, text = "E-mail: ")
    label_email.place(x = 20, y = 205)
    global entry_email
    entry_email = Entry(center_frame4, width = 28, bd = 2, textvariable = email_id)
    entry_email.place(x = 130,y = 205)
    
    
    label_gender = Label(center_frame4, text = "Gender: ")
    label_gender.place(x = 375, y = 205)
    
    global var
    var = IntVar()
    r1 = Radiobutton(center_frame4, text = "Male", padx = 10, variable = var, value = 1)
    r1.place(x = 492, y = 205)
    r2 = Radiobutton(center_frame4, text = "Female", padx = 10, variable = var, value = 2)
    r2.place(x = 550, y = 205)
    r3 = Radiobutton(center_frame4, text = "Others", padx = 10, variable = var, value = 3)
    r3.place(x = 620,y = 205)
    
    label_present_add = Label(center_frame4, text = "Present Address: ")
    label_present_add.place(x = 20 , y = 235)
    
    global entry_present_add
    entry_present_add = Entry(center_frame4, width = 28, bd = 2, textvariable = present_add)
    entry_present_add.place(x = 130,y = 235)
    
    
    label_per_add = Label(center_frame4, text = "Permanent Address: ")
    label_per_add.place(x = 375, y = 235)
    
    global entry_per_add
    entry_per_add = Entry(center_frame4, width = 28, bd = 2, textvariable = per_add)
    entry_per_add.place(x = 510,y = 235) 

    
    label_par_info = Label(center_frame4, text='Parent\'s Information',font = ("Verdana", 10, "bold"))
    label_par_info.place(x = 20, y = 265)
    
    label_fat_name = Label(center_frame4, text = "Father's Name: ")
    label_fat_name.place(x = 20, y = 295)
     
    global entry_fat_name
    entry_fat_name = Entry(center_frame4, width = 28, bd = 2, textvariable = fat_name)
    entry_fat_name.place(x = 130, y = 295)
    
    
    label_mot_name = Label(center_frame4, text = "Mother's Name: ")
    label_mot_name.place(x = 20, y = 325)
    
    global entry_mot_name
    entry_mot_name = Entry(center_frame4, width = 28, bd = 2, textvariable = mot_name)
    entry_mot_name.place(x = 130, y = 325)
    
    
    label_fat_phn = Label(center_frame4, text = "Father's Phone No.: ")
    label_fat_phn.place(x = 375, y = 295)
    
    global entry_fat_phn
    entry_fat_phn = Entry(center_frame4, width = 28, bd = 2, textvariable = fat_phn)
    entry_fat_phn.place(x = 510, y = 295)
    
    
    label_mot_phn = Label(center_frame4, text = "Mother's Phone No.: ")
    label_mot_phn.place(x = 375, y = 325)
    
    global entry_mot_phn
    entry_mot_phn = Entry(center_frame4, width = 28, bd = 2, textvariable = mot_phn)
    entry_mot_phn.place(x = 510, y = 325)
    
    
    label_class_10 = Label(center_frame4, text = "Class X",font = ("Verdana",10,"bold"))
    label_class_10.place(x = 20, y = 355)
    
    label_total_10 = Label(center_frame4, text = "Total Score: ")
    label_total_10.place(x = 20, y = 385)
    
    global entry_total_10
    entry_total_10 = Entry(center_frame4, width = 15, bd = 2, textvariable = class_10_scr)
    entry_total_10.place(x = 107, y = 385)
    
    
    label_percent_10 = Label(center_frame4, text = "Percentage: ")
    label_percent_10.place(x = 270, y = 385)
    
    global entry_percent_10
    entry_percent_10 = Entry(center_frame4, width = 15, bd = 2, textvariable = class_10_perc)
    entry_percent_10.place(x = 360 ,y = 385)
    
    
    label_board_10 = Label(center_frame4, text = "Board: ")
    label_board_10.place(x = 522, y = 385)
    label_class_12 = Label(center_frame4, text = "Class XII", font = ("Verdana",10,"bold"))
    label_class_12.place(x = 20, y = 415)
    label_total_12 = Label(center_frame4, text = "Total Score: ")
    label_total_12.place(x = 20, y = 445)
    
    global entry_total_12
    entry_total_12 = Entry(center_frame4, width = 15, bd = 2, textvariable = class_12_scr)
    entry_total_12.place(x = 107,y = 445)
    
    
    label_percent_12 = Label(center_frame4, text = "Percentage: ")
    label_percent_12.place(x = 270,y = 445)
    
    global entry_percent_12
    entry_percent_12 = Entry(center_frame4, width = 15, bd = 2, textvariable = class_12_perc)
    entry_percent_12.place(x = 360, y = 445)
    
    
    label_board_12 = Label(center_frame4, text = "Board: ")
    label_board_12.place(x = 522, y = 445)
    
    list_10 = [ 'WBBSE' , 'CBSE' , 'ICSE' , 'Other']
    
    global c1
    c1 = StringVar()   
    c1.set('Select')
    droplist_10 = OptionMenu(center_frame4, c1, *list_10)
    droplist_10.config(width = 10)
    droplist_10.place(x = 580, y = 381)
    list_12 = [ 'WBCHSE' ,'CBSE' , 'ISC' ,'Other']
    
    global c2
    c2 = StringVar()   
    c2.set('Select')
    droplist_12 = OptionMenu(center_frame4, c2, *list_12)
    droplist_12.config(width = 10)
    droplist_12.place(x = 580, y = 442)
    
   
    
    button3 = Button(center_frame4, text = 'Submit', font = ('arial', 10, 'bold'), bd = 4, bg = 'green',fg = 'white',padx = 16 ,pady = 5, command = fieldcheck, activebackground = 'light green', activeforeground = 'black')
    button3.place(x = 225, y = 485)
    button4 = Button(center_frame4, text = 'Exit', font = ('arial', 10, 'bold'),bd = 4, command = center_frame4.destroy, padx = 16, pady = 5, bg = 'black', fg = 'white', activebackground = 'silver', activeforeground = 'black')
    button4.place(x = 420, y = 485)

def frame3():
    
    center_frame3 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame3.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    
    label = Label(center_frame3, image = logo).pack(ipadx = 30, ipady = 30)

    e1 = Label(center_frame3, text = "ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font = 10)
    e1.pack(ipadx = 30)
    e2 = Label(center_frame3, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack(ipadx = 30)
    e3 = Label(center_frame3, text = 'Contact No.: (+033)2448-1081')
    e3.pack(ipadx = 30)
    e4 = Label(center_frame3, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack(ipadx = 30)
    e5 = Label(center_frame3, text = 'STUDENT REGISTRATION FORM ', fg='blue', font = 13)
    e5.pack(ipadx = 30, ipady = 10)
    
    button1 = Button(center_frame3, text = 'New Registration', command = frame4, font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button1.place(x = 175, y = 420)
    
    button2 = Button(center_frame3, text = 'Check Database', command = frame5, font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button2.place(x = 380, y = 420)
    
    button3 = Button(center_frame3, text = 'Log Out', command = center_frame3.destroy, font = ('Arial', 14), bg = 'black', fg = 'white', activebackground = 'silver', activeforeground = 'black')
    button3.place(x = 315, y = 480)

def verify_cred():
    
    global User_ID
    global Password 
    
    UID = User_ID.get()
    PD = Password.get()
    
    ID =[] 
    password =[]
    
    admin = openpyxl.load_workbook('Admin Record.xlsx')
    adminsheet = admin.active
    row = adminsheet.max_row
    column = adminsheet.max_column
    for i in range(2, row + 1):
        IDs = adminsheet.cell(row = i, column = 1)
        ID.append(str(IDs.value))
        
        passwords = adminsheet.cell(row = i, column = 2)
        password.append(str(passwords.value))
    
    if UID in ID:
        check = ID.index(UID)
        if password[check] == PD:
            User_ID.set('')
            Password.set('')
        
            frame3()
            return
        else:
            messagebox.showinfo("Mismatch", "* Office ID and Password did not match *")
            return
    else:
        messagebox.showinfo("Cannot be found", "* Office ID is not registered!! Please Sign Up *")
        return
    
def frame2():
    
    center_frame2 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame2.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    
    User_ID.set('')
    Password.set('')
  
    label = Label(center_frame2, image = logo).pack(ipadx = 30, ipady = 10)

    e1 = Label(center_frame2, text = "ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font = 10)
    e1.pack(ipadx = 30)
    e2 = Label(center_frame2, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack(ipadx = 30)
    e3 = Label(center_frame2, text = 'Contact No.: (+033)2448-1081')
    e3.pack(ipadx = 30)
    e4 = Label(center_frame2, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack(ipadx = 30)
    e5 = Label(center_frame2, text = 'STUDENT REGISTRATION FORM ', fg = 'blue', font = 13)
    e5.pack(ipadx = 30, ipady = 10)
    e6 = Label(center_frame2, text = "Enter Office ID", font = 10)
    e6.place(x = 200, y = 380)
    E1 = Entry(center_frame2, textvariable = User_ID)
    E1.place(x = 385, y = 385)
    e7 = Label(center_frame2, text = "Enter Password", font = 10)
    e7.place(x = 200, y = 410)
    E2 = Entry(center_frame2, show = '*', textvariable = Password)
    E2.place(x = 385, y = 415)
    
    button1 = Button(center_frame2, text = 'Log In', command = verify_cred, font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button1.place(x = 250, y = 460)
    
    button2 = Button(center_frame2, text = 'Back', command = center_frame2.destroy, font = ('Arial', 14), bg = 'black', fg = 'white', activebackground = 'silver', activeforeground = 'black')
    button2.place(x = 385, y = 460)
    
def signup():
    
    a = E1.get()
    b = E2.get()
    c = E3.get()
    
    if b == c:
        adminfile = openpyxl.load_workbook('Admin Record.xlsx')
        adminsheet = adminfile.active
        adminsheet.cell(column = 1,row = adminsheet.max_row + 1, value = a)
        adminsheet.cell(column = 2,row = adminsheet.max_row, value = b)
        adminfile.save('Admin Record.xlsx')
        AdminUser_ID.set('')
        AdminPassword.set('')
        ConfirmPassword.set('')
        
        center_frame1_2.destroy()
    
    else:
        messagebox.showinfo("Mismatch", "* Password did not match *")
        return

def frame1_2():
    
    global center_frame1_2
    center_frame1_2 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame1_2.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    
    AdminUser_ID.set('')
    AdminPassword.set('')
    ConfirmPassword.set('')
  
    label = Label(center_frame1_2, image = logo).pack(ipadx = 30, ipady = 10)

    e1 = Label(center_frame1_2, text = "ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font = 10)
    e1.pack(ipadx = 30)
    e2 = Label(center_frame1_2, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack(ipadx = 30)
    e3 = Label(center_frame1_2, text = 'Contact No.: (+033)2448-1081')
    e3.pack(ipadx = 30)
    e4 = Label(center_frame1_2, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack(ipadx = 30)
    e5 = Label(center_frame1_2, text = 'STUDENT REGISTRATION FORM ', fg = 'blue', font = 13)
    e5.pack(ipadx = 30, ipady = 10)
    
    e6 = Label(center_frame1_2, text = "Enter Office ID", font = 10)
    e6.place(x = 200, y = 380)
    
    global E1
    E1 = Entry(center_frame1_2, textvariable = AdminUser_ID)
    E1.place(x = 385, y = 385)
    
    e7 = Label(center_frame1_2, text = "Enter Password", font = 10)
    e7.place(x = 200, y = 403)
    
    global E2
    E2 = Entry(center_frame1_2, show = '*', textvariable = AdminPassword)
    E2.place(x = 385, y = 408)
    
    e8 = Label(center_frame1_2, text = "Re-Enter Password", font = 10)
    e8.place(x = 200, y = 426)
    
    global E3
    E3 = Entry(center_frame1_2, show = '*', textvariable = ConfirmPassword)
    E3.place(x = 385, y = 431)
    
    button1 = Button(center_frame1_2, text = 'Sign Up', command = signup, font = ('Arial', 14), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button1.place(x = 250, y = 480)
    
    button2 = Button(center_frame1_2, text = 'Back', command = center_frame1_2.destroy, font = ('Arial', 14), bg = 'black', fg = 'white', activebackground = 'silver', activeforeground = 'black')
    button2.place(x = 385, y = 480)
    
def frame1_1():
    
    center_frame1_1 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame1_1.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    
    label = Label(center_frame1_1, image = logo).pack(ipadx = 30, ipady = 30)

    e1 = Label(center_frame1_1, text = "ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font = 10)
    e1.pack(ipadx = 30)
    e2 = Label(center_frame1_1, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack(ipadx = 30)
    e3 = Label(center_frame1_1, text = 'Contact No.: (+033)2448-1081')
    e3.pack(ipadx = 30)
    e4 = Label(center_frame1_1, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack(ipadx = 30)
    e5 = Label(center_frame1_1, text = 'COLLEGE ENTRY', fg='blue', font = 13)
    e5.pack(ipadx = 30, ipady = 10)
    
    button1 = Button(center_frame1_1, text = 'Log In', command = frame2, font = ('Arial', 14), bg='SteelBlue', fg='white', activebackground='pale turquoise',activeforeground='black')
    button1.place(x = 245, y = 420)
    
    button2 = Button(center_frame1_1, text = 'Sign Up', command = frame1_2, font = ('Arial', 14), bg='SteelBlue', fg='white', activebackground='pale turquoise',activeforeground='black')
    button2.place(x = 380, y = 420)
    
    button3 = Button(center_frame1_1, text = 'Exit', command = center_frame1_1.destroy, font = ('Arial', 14), bg='black', fg='white', activebackground='silver',activeforeground='black')
    button3.place(x = 330, y = 480)
    
def frame1():
    
    center_frame1 = Frame(background_frame, relief = 'raised', borderwidth = 2)
    center_frame1.place(relx = 0.5, rely = 0.5, height = 550, width = 715, anchor = CENTER)
    label = Label(center_frame1, image = logo).pack(ipadx = 30, ipady = 30)

    e1 = Label(center_frame1, text = "ST. THOMAS' COLLEGE OF ENGINEERING AND TECHNOLOGY", font = 10)
    e1.pack(ipadx = 30)
    e2 = Label(center_frame1, text = '4, Diamond Harbour Road, Alipore Body Guard Lines, Alipore, Kolkata, West Bengal 700023') 
    e2.pack(ipadx = 30)
    e3 = Label(center_frame1, text = 'Contact No.: (+033)2448-1081')
    e3.pack(ipadx = 30)
    e4 = Label(center_frame1, text = 'Mail ID: stcet2000@gmail.com')
    e4.pack(ipadx = 30)
    e5 = Label(center_frame1, text = 'COLLEGE ENTRY', fg='blue', font = 13)
    e5.pack(ipadx = 30, ipady = 10)

    button1 = Button(center_frame1, text = 'Sign In', command = frame1_1, font = ('Arial', 15), bg = 'SteelBlue', fg = 'white', activebackground = 'pale turquoise', activeforeground = 'black')
    button1.pack(pady=10)
    button2 = Button(center_frame1, text = 'Exit', command = app.destroy, font = ('Arial', 15), bg = 'black', fg = 'white', activebackground = 'silver', activeforeground = 'black')
    button2.pack(pady = 10)
    
    
    
dim = str(GetSystemMetrics(0)) + 'x' + str(GetSystemMetrics(1))
app = Tk()
app.title('Welcome')
app.geometry(dim)

global name
global phn_no
global email_id
global present_add
global per_add
global fat_name
global fat_phn
global mot_name
global mot_phn
global class_10_scr
global class_10_perc
global class_12_scr
global class_12_perc

name = StringVar()
phn_no = StringVar()
email_id = StringVar()
present_add = StringVar()
per_add = StringVar()
fat_name = StringVar()
fat_phn = StringVar()
mot_name = StringVar()
mot_phn = StringVar()
class_10_scr = StringVar()
class_10_perc = StringVar()
class_12_scr = StringVar()
class_12_perc = StringVar()
User_ID = StringVar()
Password = StringVar()
AdminUser_ID = StringVar()
AdminPassword = StringVar()
ConfirmPassword = StringVar()

def adjust_image(event):
    new_width = event.width
    new_height = event.height
    image = image_copy.resize((new_width, new_height))
    photo = ImageTk.PhotoImage(image)
    label1.config(image = photo)
    label1.image = photo

image = Image.open('stcet image.png')
image_copy = image.copy()
photo = ImageTk.PhotoImage(image_copy)

background_frame = Frame(app, relief = 'raised', borderwidth = 2)
background_frame.pack(fill = BOTH, expand = YES)
background_frame.pack_propagate(False)

label1 = Label(background_frame, image = photo)
label1.bind('<Configure>', adjust_image)
label1.place(x = 0, y = 0, relwidth = 1, relheight=1)

logo = ImageTk.PhotoImage(Image.open("STCET_logo.png"))

frame1()

app.mainloop()

